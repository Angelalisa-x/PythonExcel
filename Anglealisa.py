from openpyxl import load_workbook
import csv, requests, re
import shutil
import os
import sys
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, BarChart3D


def create_bar_chart_1(file_path):
    """
    插入柱形图
    :param file_path: Excel 文件路径
    :return: None
    """
    wb = load_workbook(file_path)
    st = wb.active

    data1 = Reference(st, min_col=2, min_row=1, max_row=7, max_col=3)
    cats1 = Reference(st, min_col=1, min_row=2, max_row=7)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 9
    chart1.title = "日均值对比"
    # chart1.y_axis.title = '数值'
    chart1.x_axis.title = st.cell(column=1, row=1).value

    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.shape = 0
    st.add_chart(chart1, 'A8')
    wb.save(file_path)


def create_bar_chart_2(file_path):
    """
    插入3D柱形图
    :param file_path: Excel 文件路径
    :return: None
    """
    wb = load_workbook(file_path)
    st = wb.active

    data1 = Reference(st, min_col=2, min_row=1, max_row=7, max_col=3)
    cats1 = Reference(st, min_col=1, min_row=2, max_row=7)

    chart1 = BarChart3D()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "日均值对比"
    chart1.x_axis.title = None
    chart1.shape = 'cylinder'

    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    st.add_chart(chart1, 'A26')
    wb.save(file_path)

if __name__ == "__main__":
    create_bar_chart_1('lian.xlsx')
    create_bar_chart_2('lian.xlsx')